import openpyxl as o
import json


def loading(geo_file, kip_file, sklad_file):  #### Для создания json файла
    dic = {}

    prise = o.open(sklad_file, read_only=False)
    l = prise.active
    for i in range(8, l.max_row + 1):
        if l[i][0].value:
            k = l[i][0].value
            val_name = str(l[i][3].value)
            val_rrc = 'РРЦ: ' + str(l[i][5].value) + ' руб.'
            val_opt = 'ОПТ: ' + str(l[i][7].value) + ' руб.'
            val_ost_msk = 'Склад МСК: ' + str(l[i][12].value) + 'шт.'
            val_ost_rst = 'Склад РСТ: ' + str(l[i][14].value) + 'шт.'
            val_ost_krd = 'Склад КРД: ' + str(l[i][16].value) + 'шт.'
            if str(l[i][5].value) == 'None' or int(l[i][5].value) < 100:
                val_rrc = 'Уточните цену у Вашего менеджера'
            if str(l[i][7].value) == 'None':
                val_opt = 'Нет скидок на данный товар'
            if str(l[i][12].value) == 'None':
                val_ost_msk = 'МСК Нет в наличии'
            if str(l[i][14].value) == 'None':
                val_ost_rst = 'РСТ Нет в наличии'
            if str(l[i][16].value) == 'None':
                val_ost_krd = 'КРД Нет в наличии'
            dic.setdefault(k, f'{val_name}\n{val_rrc}, {val_opt}\n{val_ost_msk}, {val_ost_rst}, {val_ost_krd}')
    print('Sklad OK!')

    prise = o.open(kip_file, read_only=True)
    l = prise.worksheets[7]
    for i in range(2, l.max_row + 1):
        if l[i][2].value and l[i][5].value and l[i][6].value and l[i][0].value:
            k_kip_a = l[i][2].value.replace(' ', '').lower()
            val_kip_a_art = str(l[i][0].value)
            dic.setdefault(k_kip_a, val_kip_a_art)
    print('KIP AMO OK')

    prise = o.open(kip_file, read_only=True)
    l = prise.worksheets[3]
    for i in range(3, l.max_row + 1):
        if l[i][3].value and l[i][5].value and l[i][6].value and l[i][2].value:
            k_kip = l[i][3].value.replace('-', '').replace(' ', '').replace('с', '+').replace('.', '').lower()
            val_kip_art = str(l[i][2].value)
            dic.setdefault(k_kip, val_kip_art)
    print('KIP OK')

    prise = o.open(geo_file, read_only=False)
    l = prise.worksheets[1]
    for i in range(5, l.max_row + 1):
        if l[i][1].value and l[i][4].value and l[i][5].value and l[i][0].value:
            k_g = l[i][1].value.replace('-', '').replace(' ', '').lower()
            val_g_art = str(l[i][0].value)
            dic.setdefault(k_g, val_g_art)
    print('Geo OK!')

    l = prise.worksheets[4]
    for i in range(5, l.max_row + 1):
        if l[i][1].value and l[i][4].value and l[i][5].value and l[i][0].value:
            k_ga = l[i][1].value.replace('-', '').replace(' ', '').lower()
            val_ga_art = str(l[i][0].value)
            dic.setdefault(k_ga, val_ga_art)
    print('AMO OK!')

    with open(r'handlers\price.json', 'w', encoding='utf-8') as f:
        json.dump(dic, f, indent=2, ensure_ascii=False)
        print('json ОК!')


def test():  #### Для проверки json файла
    with open(r'handlers\price.json', 'r', encoding='utf-8') as s:
        txt = json.load(s)
    if txt['c20'] in txt:
        print(txt[txt['c20']])
    else:
        print('dic')

# loading(r"F:\RGK_priceBot\pr.xlsx", r"F:\RGK_priceBot\KIP.xlsx", r"F:\RGK_priceBot\Sklad.xlsx")   ## Запустить для создания

# test()   ### Запустить для теста
